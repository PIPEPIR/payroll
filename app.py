import streamlit as st
import pandas as pd
import io
import math
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="ระบบคิดเงินเดือนร้านอาหาร", page_icon="📝")

st.title("📝 ระบบคิดเงินเดือน")
st.write("เริ่มกะ 14.00 น. | สายไม่เกิน 14.30 หักนาทีละ 5 ฿ | สายเกิน 14.30 หักนาทีละ 10 ฿")

# --- ส่วนการตั้งค่าและอัปโหลดไฟล์ ---
if 'uploader_key' not in st.session_state:
    st.session_state.uploader_key = 0

def clear_files():
    st.session_state.uploader_key += 1

hourly_rate = st.number_input("เรทค่าจ้างต่อชั่วโมง (บาท):", min_value=1, value=50, step=5)

uploaded_files = st.file_uploader(
    "อัปโหลดไฟล์ Excel ของพนักงาน", 
    type=["xlsx"], 
    accept_multiple_files=True,
    key=f"uploader_{st.session_state.uploader_key}"
)

if st.button("🗑️ ล้างไฟล์ที่เลือกทั้งหมด", on_click=clear_files):
    pass

if uploaded_files:
    all_employees_summary = [] 
    st.divider()
    
    for file in uploaded_files:
        st.subheader(f"👤 พนักงาน/ไฟล์: {file.name}")
        
        try:
            df = pd.read_excel(file)
            df['Timestamp'] = pd.to_datetime(df['Timestamp'])
            df = df.sort_values('Timestamp').reset_index(drop=True)
            df['Date'] = df['Timestamp'].dt.date
            
            daily_records = []
            total_hours_person = 0
            total_penalty_person = 0 # เก็บยอดโดนหักรวม
            
            for date, group in df.groupby('Date'):
                punches = group['Timestamp'].tolist()
                
                if len(punches) % 2 != 0:
                    st.warning(f"⚠️ วันที่ {date}: มีการตอกบัตร {len(punches)} ครั้ง ระบบจะคิดเฉพาะคู่ที่สมบูรณ์")
                
                # ==========================================
                # ระบบคำนวณหักเงินมาสาย (ดูจากการตอกบัตรรอบแรกของวัน)
                # ==========================================
                first_punch = punches[0]
                shift_start_time = first_punch.replace(hour=14, minute=0, second=0, microsecond=0)
                
                daily_penalty = 0
                late_mins = 0
                
                # ถ้าตอกบัตรเข้างานหลัง 14:00 น.
                if first_punch > shift_start_time:
                    late_delta = first_punch - shift_start_time
                    # ปัดเศษนาทีลง (ถ้ามา 14:00:59 ถือว่าไม่สาย)
                    late_mins = math.floor(late_delta.total_seconds() / 60) 
                    
                    if late_mins > 0:
                        if late_mins <= 30:
                            daily_penalty = late_mins * 5
                        else:
                            # 30 นาทีแรก นาทีละ 5 บาท + นาทีที่เกิน 30 นาทีละ 10 บาท
                            daily_penalty = (30 * 5) + ((late_mins - 30) * 10)
                
                total_penalty_person += daily_penalty

                # ==========================================
                # คำนวณชั่วโมงทำงานปกติ
                # ==========================================
                daily_hours = 0
                for i in range(0, len(punches) - 1, 2):
                    time_in = punches[i]
                    time_out = punches[i+1]
                    hours = (time_out - time_in).total_seconds() / 3600
                    daily_hours += hours
                
                daily_hours = round(daily_hours, 2)
                total_hours_person += daily_hours
                
                daily_records.append({
                    'วันที่': date,
                    'เวลาเข้างาน (รอบแรก)': first_punch.strftime('%H:%M:%S'),
                    'สาย (นาที)': late_mins,
                    'โดนหัก (บาท)': daily_penalty,
                    'ชั่วโมงทำงานรวม': daily_hours
                })
            
            # แสดงตารางรายวัน
            if daily_records:
                with st.expander(f"ดูรายละเอียดรายวัน ของ {file.name}"):
                    st.dataframe(pd.DataFrame(daily_records))
            
            # สรุปยอดเงินของคนนี้
            base_pay = total_hours_person * hourly_rate
            net_pay = base_pay - total_penalty_person
            
            st.success(f"ทำงาน: {total_hours_person:.2f} ชม. | ค่าจ้าง: ฿{base_pay:,.2f} | โดนหักสาย: ฿{total_penalty_person:,.2f} | **รับสุทธิ: ฿{net_pay:,.2f}**")
            st.write("---")
            
            all_employees_summary.append({
                "ชื่อไฟล์ (พนักงาน)": file.name,
                "ชั่วโมงทำงาน (ชม.)": total_hours_person,
                "ค่าจ้างปกติ (บาท)": base_pay,
                "หักมาสาย (บาท)": total_penalty_person,
                "รับเงินสุทธิ (บาท)": net_pay
            })
            
        except Exception as e:
            st.error(f"ไฟล์ {file.name} มีปัญหา (Error: {e})")

    # สรุปยอดรวมทุกคน
    if all_employees_summary:
        st.header("💰 สรุปยอดจ่ายเงินพนักงานทั้งหมด")
        summary_df = pd.DataFrame(all_employees_summary)
        st.dataframe(summary_df, use_container_width=True)
        
        grand_total = summary_df['รับเงินสุทธิ (บาท)'].sum()
        st.metric("ยอดเงินรวมที่ร้านต้องโอนจ่าย (บาท)", f"฿{grand_total:,.2f}")

        # ==========================================
        # ฟีเจอร์ Export เป็น Excel แบบจัดรูปแบบสวยงาม
        # ==========================================
        def to_excel(df):
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='สรุปยอดจ่ายเงิน')
                
                # เข้าถึง workbook และ worksheet
                workbook = writer.book
                worksheet = writer.sheets['สรุปยอดจ่ายเงิน']

                # --- 1. กำหนดสไตล์ ---
                header_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF", size=12)
                alignment_center = Alignment(horizontal="center", vertical="center")
                border_thin = Border(
                    left=Side(style='thin'), 
                    right=Side(style='thin'), 
                    top=Side(style='thin'), 
                    bottom=Side(style='thin')
                )

                # --- 2. จัดรูปแบบ Header ---
                for col_num, column_title in enumerate(df.columns, 1):
                    cell = worksheet.cell(row=1, column=col_num)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = alignment_center
                    cell.border = border_thin

                # --- 3. ปรับความกว้างคอลัมน์อัตโนมัติและใส่เส้นขอบ ---
                for col_num, column_title in enumerate(df.columns, 1):
                    max_length = 0
                    column_letter = get_column_letter(col_num)
                    
                    # คำนวณความยาวสูงสุดในคอลัมน์นั้นๆ
                    column_cells = worksheet[column_letter]
                    for cell in column_cells:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                        
                        # ใส่เส้นขอบทุกเซลล์
                        cell.border = border_thin
                        # ถ้าเป็นตัวเลข ให้จัดกลาง (ยกเว้นชื่อพนักงาน)
                        if col_num > 1:
                            cell.alignment = alignment_center

                    adjusted_width = (max_length + 5)
                    worksheet.column_dimensions[column_letter].width = adjusted_width

            processed_data = output.getvalue()
            return processed_data

        excel_data = to_excel(summary_df)
        
        st.download_button(
            label="📥 ดาวน์โหลดสรุปยอดเงินทั้งหมด (Excel)",
            data=excel_data,
            file_name=f"payroll_summary_{pd.Timestamp.now().strftime('%Y%m%d_%H%M')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )